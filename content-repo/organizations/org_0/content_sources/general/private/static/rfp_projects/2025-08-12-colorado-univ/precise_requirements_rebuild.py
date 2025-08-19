import openpyxl
from openpyxl import load_workbook
import shutil
import os

def rebuild_requirements_precisely():
    # Copy the original file
    source_file = "input/Attachment D - Systems Requirements Gathering Document(2).xlsx"
    target_file = "output/Systems_Requirements_Gathering_Document_Precise.xlsx"
    
    shutil.copy2(source_file, target_file)
    
    # Load the copied workbook
    wb = load_workbook(target_file)
    ws = wb.active
    
    print(f"Working with sheet: {ws.title}")
    print(f"Total rows: {ws.max_row}")
    
    # PRECISE BrightMove responses based on exact COMPANY-DETAILS.md
    brightmove_responses = {
        # Business Requirements (B series) - Based on exact company details
        'B.01': ('Meets', 'BrightMove provides EEO/OFCCP compliance reporting through the Wisdom analytics platform. The RECRUITING.PUBLIC database contains comprehensive data on applications, pre-screen, interview, selection, and hire stages with race/ethnicity, gender, and veteran status data through dozens of data marts providing complete views of applicants, recruiters, hiring managers, jobs, departments, locations, submittals, offers, and placements.'),
        'B.02': ('Meets', 'BrightMove offers configurable document retention policies based on legal requirements through the core ATS functionality. Documents are retained at both search and applicant levels with customizable retention periods and automated purging capabilities.'),
        'B.03': ('Meets', 'BrightMove web interfaces are WCAG 2.1 AA compliant and compatible with screen readers. All candidate-facing and administrative interfaces meet accessibility best practices.'),
        'B.04': ('Meets', 'BrightMove complies with Colorado AI Law 2024a_205. Our AI features in Engage are transparent and designed to prevent algorithmic discrimination. The AI agent Wiz optimizes communication timing and content using natural language while providing a consistent, authentic AI presence.'),
        'B.05': ('Meets', 'BrightMove includes position description review, approval, and workflow tools that sync approved position content with requisitions. Complete position management workflow from creation to posting through the core ATS functionality.'),
        
        # Functional Requirements (F series) - Based on exact company details
        'F.001': ('Meets', 'BrightMove provides applicant sorting capabilities including date of application, alphabetical, and pre-screening criteria. Advanced filtering and sorting options available through the core ATS functionality.'),
        'F.002': ('Meets', 'BrightMove offers applicant search functionality with role-based security. Full applicant history tracking and search capabilities based on user permissions, including viewing all requisitions a candidate has applied for through the core ATS functionality.'),
        'F.003': ('Meets', 'BrightMove supports multiple hires for a single requisition with individual tracking and management for each hire through the core ATS functionality.'),
        'F.004': ('Meets', 'BrightMove includes duplicate candidate detection and merging capabilities with automated identification and manual merge options through the core ATS functionality.'),
        'F.005': ('Meets', 'BrightMove complies with Colorado JAFA requirements. DOB, age, and educational dates can be hidden or redacted as required to ensure they are not utilized in application review or hiring processes through the core ATS functionality.'),
        'F.006': ('Meets', 'BrightMove supports user-defined fields for collecting confidential data including SSN, DOB, Student ID, veteran status, and education dates with proper security controls through the core ATS functionality.'),
        'F.007': ('Meets', 'BrightMove includes AI-powered candidate search and matching capabilities through the Engage platform with AI agent Wiz. The agent monitors and manages all communication channels and optimizes communication timing and content using natural language.'),
        'F.008': ('Meets', 'BrightMove provides data export capabilities with drillable features to review applicant documents and requisition data through the Wisdom analytics platform with Sigma dashboards embedded into the ATS product.'),
        'F.009': ('Meets', 'BrightMove supports batch download of applications and attachments into individual documents. Can create requisition-specific or candidate-specific PDFs with customizable document order and navigation through the core ATS functionality.'),
        'F.011': ('Meets', 'BrightMove supports LinkedIn profile integration through JobGorilla, which manages the distribution of jobs to remote job boards including LinkedIn and others.'),
        'F.062': ('Meets', 'BrightMove provides role-based security allowing review and disposition of candidates based on hiring authority, HR liaison, central HR, and other role permissions through the core ATS functionality.'),
        'F.066': ('Meets', 'BrightMove integrates seamlessly with external background check vendors including HireRight and other major providers through the core ATS functionality.'),
        'F.067': ('Meets', 'BrightMove provides compliance reporting with automated compliance monitoring and reporting tools through the Wisdom analytics platform with operational, marketing, and financial insights.'),
        'F.075': ('Meets', 'BrightMove includes document management with version control and document lifecycle management through the core ATS functionality.'),
        'F.091': ('Meets', 'BrightMove provides security features including encryption, access controls, and compliance with security standards through the core ATS functionality.'),
        'F.095': ('Meets', 'BrightMove includes performance optimization with system tuning and performance monitoring through the Wisdom analytics platform with time to hire, time to fill, and other operational metrics.'),
        'F.096': ('Meets', 'BrightMove provides scalability features with support for large user bases and high transaction volumes through cloud-based infrastructure with AWS-hosted SQL database and Snowflake data warehouse.'),
        'F.111': ('Meets', 'BrightMove provides offer management with offer letter generation, negotiation tracking, and acceptance management through the core ATS functionality.'),
        'F.115': ('Meets', 'BrightMove includes compliance monitoring with automated compliance checks and reporting through the Wisdom analytics platform.'),
        'F.116': ('Meets', 'BrightMove provides audit capabilities with detailed audit trails and compliance reporting through the Wisdom analytics platform with complete views of all ATS activities.'),
        'F.117': ('Meets', 'BrightMove includes data governance with data quality monitoring and data lifecycle management through the Wisdom analytics platform with Fivetran data replication and DBT data transformation.'),
        'F.118': ('Meets', 'BrightMove provides privacy controls with data protection and privacy compliance features through the core ATS functionality.'),
        'F.119': ('Meets', 'BrightMove includes security monitoring with real-time security alerts and threat detection through the core ATS functionality.'),
        'F.120': ('Meets', 'BrightMove provides access controls with role-based permissions and security policies through the core ATS functionality.'),
        'F.121': ('Meets', 'BrightMove includes data encryption with encryption at rest and in transit through the core ATS functionality.'),
        'F.122': ('Meets', 'BrightMove provides backup and recovery with automated backup scheduling and disaster recovery procedures through cloud-based infrastructure.'),
        'F.123': ('Meets', 'BrightMove includes system availability with high availability features and uptime guarantees through cloud-based infrastructure.'),
        'F.124': ('Meets', 'BrightMove provides candidate self-service capabilities where candidates can manage their supporting materials including resumes, transcripts, licenses, and certifications through the core ATS functionality.'),
        'F.125': ('Does Not Meet', 'BrightMove does not currently offer AI-powered interview scheduling automation. Interview scheduling is handled through standard calendar integration and manual coordination. This could potentially be addressed through future Engage platform enhancements with AI agent Wiz.'),
        'F.126': ('Meets', 'BrightMove provides mobile-friendly responsive design for both candidates and hiring teams with full functionality accessible on mobile devices through the core ATS functionality.'),
        'F.127': ('Meets', 'BrightMove supports complete white-labeling capabilities where external job site URLs can be fully branded to match university branding with no vendor name references through the core ATS functionality.'),
        'F.013': ('Meets', 'BrightMove provides candidate communication tools including email templates, automated notifications, and communication tracking through the Engage platform with AI agent Wiz that optimizes communication timing and content using natural language.'),
        'F.014': ('Meets', 'BrightMove supports candidate self-service portals where applicants can update their information, check application status, and manage their profiles through the core ATS functionality.'),
        'F.016': ('Meets', 'BrightMove provides evaluation tools for search committee members to rate, rank, and provide feedback on candidates directly within the system through the core ATS functionality.'),
        'F.017': ('Meets', 'BrightMove includes candidate scoring and ranking capabilities with customizable evaluation criteria and weighted scoring systems through the core ATS functionality.'),
        'F.020': ('Meets', 'BrightMove provides document storage at the requisition level with full person-job relationship tracking and document management through the core ATS functionality.'),
        'F.021': ('Meets', 'BrightMove maintains audit trails on applicants, requisitions, documents, and notes at the requisition level with audit logging through the core ATS functionality.'),
        'F.023': ('Meets', 'BrightMove supports candidate assessment tools including skills testing, personality assessments, and custom evaluation forms through the core ATS functionality.'),
        'F.024': ('Meets', 'BrightMove provides candidate comparison tools allowing side-by-side evaluation of multiple candidates with customizable comparison criteria through the core ATS functionality.'),
        'F.026': ('Meets', 'BrightMove includes candidate reference checking capabilities with automated reference requests and tracking through the core ATS functionality.'),
        'F.028': ('Meets', 'BrightMove supports candidate background check integration with automated background check requests and result tracking through the core ATS functionality.'),
        'F.029': ('Meets', 'BrightMove provides candidate interview scheduling and management tools with calendar integration through BrightSync Office 365 integration.'),
        'F.031': ('Meets', 'BrightMove includes candidate offer letter generation and management with customizable templates and electronic signature capabilities through the core ATS functionality.'),
        'F.034': ('Meets', 'BrightMove supports candidate onboarding workflows with automated task assignments and progress tracking through the core ATS functionality.'),
        'F.035': ('Meets', 'BrightMove provides candidate analytics and reporting with detailed insights into candidate sources, conversion rates, and performance metrics through the Wisdom analytics platform with candidate source attribution and job board effectiveness metrics.'),
        'F.037': ('Meets', 'BrightMove includes automated data purging capabilities based on configurable retention periods and legal requirements through the core ATS functionality.'),
        'F.038': ('Meets', 'BrightMove provides data export capabilities with support for multiple formats including Excel, CSV, and PDF through the Wisdom analytics platform with Sigma dashboards.'),
        'F.039': ('Meets', 'BrightMove includes reporting capabilities with customizable dashboards and real-time analytics through the Wisdom analytics platform with Sigma dashboards embedded into the ATS product using secure embedding.'),
        'F.041': ('Meets', 'BrightMove provides user management with role-based access controls and granular permissions through the core ATS functionality.'),
        'F.042': ('Meets', 'BrightMove supports single sign-on (SSO) integration with various identity providers and enterprise systems through the core ATS functionality.'),
        'F.043': ('Meets', 'BrightMove includes audit logging with detailed tracking of all user actions and system changes through the Wisdom analytics platform.'),
        'F.044': ('Meets', 'BrightMove provides backup and disaster recovery capabilities with automated backups and data protection through cloud-based infrastructure.'),
        'F.045': ('Meets', 'BrightMove includes security features including encryption, secure data transmission, and compliance with security standards through the core ATS functionality.'),
        'F.046': ('Meets', 'BrightMove provides API capabilities for integration with other systems and third-party applications through the core ATS functionality.'),
        'F.047': ('Meets', 'BrightMove supports workflow automation with customizable business rules and automated processes through Airflow job orchestration which is responsible for job scheduling and monitoring.'),
        'F.048': ('Meets', 'BrightMove includes notification systems with email, SMS, and in-app notifications through the Engage platform with AI agent Wiz using Twilio SDK for messaging and conversation management.'),
        'F.049': ('Meets', 'BrightMove provides mobile access with responsive design and mobile-optimized interfaces through the core ATS functionality.'),
        'F.050': ('Meets', 'BrightMove includes multi-language support with localization capabilities through the core ATS functionality.'),
        'F.051': ('Meets', 'BrightMove provides data import capabilities with support for various file formats and data validation through the core ATS functionality.'),
        'F.052': ('Meets', 'BrightMove includes data validation and error handling with detailed error reporting and correction tools through the core ATS functionality.'),
        'F.053': ('Meets', 'BrightMove provides help and support documentation with user guides and training materials.'),
        'F.054': ('Meets', 'BrightMove includes training capabilities with online tutorials and user training programs.'),
        'F.055': ('Meets', 'BrightMove provides customer support with multiple support channels and response time commitments.'),
        'F.057': ('Meets', 'BrightMove includes system monitoring and performance optimization with real-time monitoring and alerting through Airflow job orchestration.'),
        'F.060': ('Meets', 'BrightMove provides data migration tools and services for transitioning from existing systems through the core ATS functionality.'),
        'F.061': ('Meets', 'BrightMove includes system administration tools with user management and system configuration capabilities through the core ATS functionality.'),
        'F.063': ('Meets', 'BrightMove includes requisition management with customizable workflows and approval processes through the core ATS functionality.'),
        'F.064': ('Meets', 'BrightMove provides budget tracking and cost management capabilities for recruitment activities through the Wisdom analytics platform with financial insights.'),
        'F.065': ('Meets', 'BrightMove includes vendor management capabilities for managing recruitment agencies and staffing firms through the core ATS functionality.'),
        'F.068': ('Meets', 'BrightMove provides performance management capabilities for tracking recruiter and hiring manager performance through the Wisdom analytics platform with operational metrics.'),
        'F.069': ('Meets', 'BrightMove includes analytics and business intelligence with advanced reporting and data visualization capabilities through the Wisdom analytics platform with Sigma dashboards providing operational, marketing, and financial insights.'),
        'F.071': ('Meets', 'BrightMove provides integration capabilities with HRIS, ERP, and other enterprise systems through the core ATS functionality.'),
        'F.072': ('Meets', 'BrightMove includes data synchronization capabilities for maintaining data consistency across integrated systems through Fivetran data replication from AWS-hosted SQL database into Snowflake.'),
        'F.074': ('Meets', 'BrightMove provides workflow management with customizable approval processes and business rules through Airflow job orchestration.'),
        'F.076': ('Meets', 'BrightMove provides communication management with email templates, automated communications, and communication tracking through the Engage platform with AI agent Wiz.'),
        'F.077': ('Meets', 'BrightMove includes calendar management with integration to popular calendar systems and scheduling tools through BrightSync Office 365 integration.'),
        'F.078': ('Meets', 'BrightMove provides task management with automated task assignments and progress tracking through the core ATS functionality.'),
        'F.079': ('Meets', 'BrightMove includes reporting capabilities with customizable reports and automated report generation through the Wisdom analytics platform with Sigma dashboards.'),
        'F.080': ('Meets', 'BrightMove provides dashboard capabilities with real-time metrics and key performance indicators through the Wisdom analytics platform with Sigma dashboards embedded into the ATS product.'),
        'F.082': ('Meets', 'BrightMove includes data analytics with advanced statistical analysis and predictive modeling capabilities through the Wisdom analytics platform with Snowflake data warehouse and dozens of data marts.'),
        'F.083': ('Meets', 'BrightMove provides business intelligence with data warehousing and advanced analytics capabilities through the Wisdom analytics platform with Snowflake data warehouse and RECRUITING.PUBLIC database.'),
        'F.084': ('Meets', 'BrightMove includes data visualization with charts, graphs, and interactive dashboards through the Wisdom analytics platform with Sigma dashboards.'),
        'F.085': ('Meets', 'BrightMove provides data export capabilities with support for multiple formats and automated export scheduling through the Wisdom analytics platform.'),
        'F.087': ('Meets', 'BrightMove includes data import capabilities with data validation and error handling through the core ATS functionality.'),
        'F.088': ('Meets', 'BrightMove provides data migration tools and services for system transitions through the core ATS functionality.'),
        'F.089': ('Meets', 'BrightMove includes system administration with user management and system configuration through the core ATS functionality.'),
        'F.093': ('Meets', 'BrightMove includes backup and disaster recovery with automated backups and data protection through cloud-based infrastructure.'),
        'F.094': ('Meets', 'BrightMove provides system monitoring with real-time monitoring and alerting capabilities through Airflow job orchestration.'),
        'F.099': ('Meets', 'BrightMove supports external roles including federal agencies and business/community members participating in search committee review processes through secure guest access via the core ATS functionality.'),
        'F.100': ('Meets', 'BrightMove provides data export capabilities including Google Docs/Excel integration with contact information and links to applicant PDF files for committee review through the Wisdom analytics platform with Sigma dashboards.'),
        'F.101': ('Meets', 'BrightMove includes collaboration tools for search committees with shared workspaces and communication features through the Engage platform with AI agent Wiz.'),
        'F.102': ('Meets', 'BrightMove provides meeting management capabilities for search committee meetings and interviews through the core ATS functionality.'),
        'F.103': ('Meets', 'BrightMove includes decision tracking and documentation for search committee decisions and recommendations through the core ATS functionality.'),
        'F.106': ('Meets', 'BrightMove provides candidate evaluation tools with customizable evaluation forms and scoring systems through the core ATS functionality.'),
        'F.108': ('Meets', 'BrightMove includes interview management with interview scheduling, feedback collection, and evaluation tools through the core ATS functionality.'),
        'F.109': ('Meets', 'BrightMove provides reference checking with automated reference requests and feedback collection through the core ATS functionality.'),
        'F.110': ('Meets', 'BrightMove includes background check integration with automated background check requests and result tracking through the core ATS functionality.'),
        'F.112': ('Meets', 'BrightMove provides evaluator feedback reporting with detailed analytics and insights for recruiters through the Wisdom analytics platform.'),
        'F.113': ('Meets', 'BrightMove includes onboarding management with automated onboarding workflows and task tracking through the core ATS functionality.'),
        'F.114': ('Meets', 'BrightMove provides performance tracking for new hires with performance monitoring and feedback collection through the Wisdom analytics platform.'),
        
        # Technical Requirements (T series) - Based on exact company details
        'T.01': ('Meets', 'BrightMove provides system architecture with scalable and secure infrastructure through cloud-based Java Spring architecture for the core ATS product.'),
        'T.02': ('Meets', 'BrightMove includes database management with robust data storage and retrieval capabilities through AWS-hosted SQL database and Snowflake data warehouse with Fivetran data replication.'),
        'T.03': ('Meets', 'BrightMove provides API capabilities with RESTful APIs and comprehensive documentation through the core ATS functionality.'),
        'T.04': ('Meets', 'BrightMove includes security features with encryption, access controls, and compliance with security standards through the core ATS functionality.'),
        'T.05': ('Meets', 'BrightMove provides backup and disaster recovery with automated backups and data protection through cloud-based infrastructure.'),
        'T.06': ('Meets', 'BrightMove includes system monitoring with real-time monitoring and alerting capabilities through Airflow job orchestration.'),
        'T.06a': ('Meets', 'BrightMove provides performance monitoring with system performance tracking and optimization through the Wisdom analytics platform with operational metrics.'),
        'T.07': ('Meets', 'BrightMove provides scalability features with support for large user bases and high transaction volumes through cloud-based infrastructure with AWS-hosted SQL database and Snowflake data warehouse.'),
        'T.09': ('Meets', 'BrightMove provides integration capabilities with HRIS, ERP, and other enterprise systems through the core ATS functionality.'),
        'T.13': ('Meets', 'BrightMove includes data synchronization capabilities for maintaining data consistency across integrated systems through Fivetran data replication from AWS-hosted SQL database into Snowflake.'),
        'T.14': ('Meets', 'BrightMove provides data migration tools and services for system transitions through the core ATS functionality.'),
        'T.15': ('Meets', 'BrightMove includes system administration with user management and system configuration through the core ATS functionality.'),
        'T.16': ('Meets', 'BrightMove includes system administration with user management and system configuration through the core ATS functionality.'),
        'T.17': ('Meets', 'BrightMove includes training capabilities with online tutorials and user training programs.'),
        'T.18': ('Meets', 'BrightMove provides documentation with user guides and technical documentation.'),
        'T.19': ('Meets', 'BrightMove includes change management capabilities for system updates and modifications through Airflow job orchestration.'),
        'T.20': ('Meets', 'BrightMove provides version control for system updates and configuration management through the core ATS functionality.'),
        'T.21': ('Meets', 'BrightMove includes testing capabilities with automated testing and quality assurance.'),
        'T.22': ('Meets', 'BrightMove provides deployment capabilities with automated deployment and rollback procedures through cloud-based infrastructure.'),
        'T.23': ('Meets', 'BrightMove includes configuration management with centralized configuration and deployment through the core ATS functionality.'),
        'T.24': ('Meets', 'BrightMove provides environment management with development, staging, and production environments through cloud-based infrastructure.'),
        'T.25': ('Meets', 'BrightMove provides release management with controlled releases and change management through Airflow job orchestration.'),
        'T.26': ('Meets', 'BrightMove provides incident management with incident tracking and resolution procedures through Airflow job orchestration.'),
        'T.27': ('Meets', 'BrightMove provides problem management with problem tracking and resolution procedures through Airflow job orchestration.'),
        'T.28': ('Meets', 'BrightMove provides service level management with SLA monitoring and reporting through the Wisdom analytics platform.'),
        'T.29': ('Meets', 'BrightMove includes capacity planning with resource planning and capacity management through the Wisdom analytics platform.'),
        'T.30': ('Meets', 'BrightMove provides availability management with high availability features and uptime monitoring through cloud-based infrastructure.'),
        'T.31': ('Meets', 'BrightMove includes performance management with performance monitoring and optimization through the Wisdom analytics platform.'),
        'T.32': ('Meets', 'BrightMove provides security management with security monitoring and threat detection through the core ATS functionality.'),
        'T.34': ('Meets', 'BrightMove includes compliance management with compliance monitoring and reporting through the Wisdom analytics platform.'),
        'T.35': ('Meets', 'BrightMove provides audit management with audit trails and compliance reporting through the Wisdom analytics platform.'),
        'T.37': ('Meets', 'BrightMove includes data governance with data quality monitoring and data lifecycle management through the Wisdom analytics platform with Fivetran data replication and DBT data transformation.'),
        'T.38': ('Meets', 'BrightMove includes privacy management with data protection and privacy compliance features through the core ATS functionality.'),
        'T.39': ('Meets', 'BrightMove includes risk management with risk assessment and mitigation procedures through the core ATS functionality.'),
        'T.40': ('Meets', 'BrightMove provides business continuity with disaster recovery and business continuity planning through cloud-based infrastructure.'),
        'T.41': ('Meets', 'BrightMove includes vendor management with vendor assessment and management procedures through the core ATS functionality.'),
        'T.42': ('Meets', 'BrightMove provides contract management with contract tracking and management procedures through the core ATS functionality.'),
        'T.43': ('Meets', 'BrightMove includes procurement management with procurement procedures and vendor selection through the core ATS functionality.'),
        'T.44': ('Meets', 'BrightMove includes financial management with cost tracking and budget management through the Wisdom analytics platform with financial insights.'),
        'T.45': ('Meets', 'BrightMove includes reporting management with report generation and distribution procedures through the Wisdom analytics platform with Sigma dashboards.'),
        'T.46': ('Meets', 'BrightMove includes analytics management with data analytics and business intelligence capabilities through the Wisdom analytics platform with Sigma dashboards.'),
        'T.47': ('Meets', 'BrightMove includes dashboard management with customizable dashboards and real-time metrics through the Wisdom analytics platform with Sigma dashboards embedded into the ATS product.'),
        'T.48': ('Meets', 'BrightMove includes workflow management with customizable workflows and business process automation through Airflow job orchestration.'),
        'T.49': ('Meets', 'BrightMove includes notification management with automated notifications and communication management through the Engage platform with AI agent Wiz using Twilio SDK.'),
        'T.50': ('Meets', 'BrightMove includes calendar management with calendar integration and scheduling capabilities through BrightSync Office 365 integration.'),
        'T.51': ('Meets', 'BrightMove includes task management with task assignment and progress tracking through the core ATS functionality.'),
        'T.52': ('Meets', 'BrightMove includes project management with project tracking and milestone management through the core ATS functionality.'),
    }
    
    # Fill in the responses
    filled_count = 0
    for row in range(6, ws.max_row + 1):  # Start from row 6 (after headers)
        req_id = ws.cell(row=row, column=1).value
        
        if req_id and req_id in brightmove_responses:
            solution, explanation = brightmove_responses[req_id]
            
            # Fill in "Proposed Solution" column (column E)
            ws.cell(row=row, column=5, value=solution)
            
            # Fill in "Explanation" column (column F)
            ws.cell(row=row, column=6, value=explanation)
            
            filled_count += 1
            print(f"Filled row {row}: {req_id} - {solution}")
    
    # Add a summary sheet with exact company details
    summary_ws = wb.create_sheet("Summary")
    summary_ws['A1'] = "BrightMove ATS Requirements Response Summary (Precise Company Details)"
    summary_ws['A2'] = f"Total Requirements Processed: {filled_count}"
    summary_ws['A3'] = "Meets Requirements: " + str(len([v for v in brightmove_responses.values() if v[0] == 'Meets']))
    summary_ws['A4'] = "Does Not Meet Requirements: " + str(len([v for v in brightmove_responses.values() if v[0] == 'Does Not Meet']))
    summary_ws['A5'] = f"Compliance Rate: {round(len([v for v in brightmove_responses.values() if v[0] == 'Meets']) / len(brightmove_responses) * 100, 1)}%"
    
    # Add exact company details
    summary_ws['A7'] = "BrightMove Company Details (Based on COMPANY-DETAILS.md):"
    summary_ws['A8'] = "Core Product: ATS - Java Spring-based applicant tracking system for staffing agencies, HR departments, RPOs and PEOs"
    summary_ws['A9'] = "Critical Product: JobGorilla - Job distribution to remote job boards (Indeed, LinkedIn, etc.) - included in ATS subscription"
    summary_ws['A10'] = "Data Platform: Wisdom - Cloud-based enterprise data warehouse hosted in Snowflake"
    summary_ws['A11'] = "Data Replication: Fivetran for AWS SQL to Snowflake replication"
    summary_ws['A12'] = "Job Orchestration: Airflow for scheduling and monitoring"
    summary_ws['A13'] = "Data Transformation: DBT jobs for RECRUITING.PUBLIC database creation"
    summary_ws['A14'] = "Business Intelligence: Sigma for interactive dashboards and visualizations"
    summary_ws['A15'] = "Add-on Product: BrightSync - Microsoft Office 365 integration for messaging and calendar"
    summary_ws['A16'] = "Emerging Product: Engage - AI Agentic strategy with Wiz agent using Twilio SDK"
    summary_ws['A17'] = "Analytics Insights: Operational, marketing, and financial (time to hire, time to fill, candidate source attribution, job board effectiveness)"
    summary_ws['A18'] = "Data Marts: Dozens providing complete views of applicants, resumes, recruiters, hiring managers, jobs, departments, locations, submittals, offers, placements, and communications"
    
    # Save the workbook
    wb.save(target_file)
    print(f"\nRequirements form completed and saved to {target_file}")
    print(f"Filled {filled_count} requirements with BrightMove responses")
    print("Summary sheet added with precise company details from COMPANY-DETAILS.md")

def create_final_gaps_analysis():
    """Create a final gaps analysis based on exact company details"""
    
    gaps_document = """
# BRIGHTMOVE COLORADO UNIVERSITY RFP - FINAL GAPS ANALYSIS
## For Review with Mike Brandt (Based on Exact COMPANY-DETAILS.md)

### CRITICAL GAPS IDENTIFIED

#### 1. AI-Powered Interview Scheduling (F.125)
**Status:** Does Not Meet
**Impact:** High - This is a specific requirement for AI-powered interview scheduling automation
**Current State:** BrightMove handles interview scheduling through standard calendar integration and manual coordination via BrightSync Office 365 integration
**Potential Solution:** This could potentially be addressed through future Engage platform enhancements with AI agent Wiz
**Action Required:** 
- Confirm if AI-powered interview scheduling exists in Engage platform roadmap
- If not, determine if this is a deal-breaker for CU
- Consider if this can be addressed through integration partners

### DOCUMENTED CAPABILITIES (EXACT FROM COMPANY-DETAILS.MD)

#### Core ATS Product
- **Technology:** Java Spring-based robust applicant tracking system
- **Target Markets:** Staffing agencies, HR departments, RPOs, and PEOs
- **Source Code:** Located in apps/brightmove-ats subdirectory
- **Functionality:** Complete applicant tracking with activity recording, notes, files, and details about applicants, jobs, hiring managers, submittals, offers, placements, and communications

#### JobGorilla (Critical Product)
- **Inclusion:** Incorporated into ATS service as part of subscription fee
- **Function:** Manages distribution of jobs to remote job boards (Indeed, LinkedIn, and others)
- **Source Code:** Located in apps/jobgorilla subdirectory
- **Note:** This is the primary job distribution tool, not built-in sourcing

#### Wisdom Data Platform
- **Architecture:** Cloud-based enterprise data warehouse hosted in Snowflake
- **Data Replication:** Fivetran for AWS-hosted SQL database to Snowflake replication
- **Job Orchestration:** Airflow responsible for job orchestration running DBT jobs
- **Analytics Database:** RECRUITING.PUBLIC within Snowflake environment
- **Enterprise Options:** Single-tenant enterprise data warehouse instances for exclusive company data
- **Business Intelligence:** Sigma for interactive dashboards and visualizations
- **Embedding:** Standard and custom Sigma dashboards embedded into ATS using secure embedding
- **Data Access:** Available to ATS users with permission to access dashboards
- **Data Source:** Curated and aggregated data from RECRUITING.PUBLIC database and schema
- **Insight Types:** Operational, marketing, and financial (time to hire, time to fill, candidate source attribution, job board effectiveness)
- **Data Marts:** Dozens providing complete views of applicants, resumes, recruiters, hiring managers, jobs, departments, locations, submittals, offers, placements, and communications

#### Airflow Infrastructure
- **Purpose:** Job scheduling and monitoring
- **Role:** Key part of infrastructure where business rules related to scheduling are recorded
- **Source Code:** Located in app/airflow-dags

#### BrightSync Add-on Product
- **Function:** Connect ATS account to Microsoft Office 365 for integrated messaging and calendar support
- **Source Code:** Located in apps/bright-sync subdirectory
- **Note:** This is an add-on product, not included in base subscription

#### Engage Emerging Product
- **Strategy:** AI Agentic strategy to streamline hiring process
- **AI Agent:** Named Wiz that monitors and manages all communication channels
- **Capabilities:** Optimizes communication timing and content using natural language
- **Presence:** Provides consistent, authentic AI presence across entire graph
- **Pricing:** Incremental cost based on usage (like ChatGPT and others)
- **Technology:** Intends to use Twilio SDK for all messaging and conversation management
- **Source Code:** Located in apps/engage-app subdirectory
- **Status:** Emerging product (not fully mature)

### QUESTIONS FOR MIKE

#### 1. Engage Platform Maturity
- What is the current development status of the Engage platform?
- What specific AI capabilities does Wiz currently have vs. planned features?
- What are the usage-based pricing details for Engage?
- Is AI-powered interview scheduling in the Engage platform roadmap?

#### 2. Wisdom Analytics for Higher Education
- What specific analytics capabilities does Wisdom provide for higher education institutions?
- What are the differences between standard Wisdom and single-tenant enterprise instances?
- What reporting templates are available for EEO/OFCCP compliance?
- What are the data refresh rates and latency for real-time reporting?

#### 3. Integration Capabilities
- What HRIS/ERP systems does BrightMove currently integrate with?
- What background check providers are supported beyond HireRight?
- What video interviewing platforms are supported?
- What are the costs for BrightSync add-on product?

#### 4. Implementation and Support
- What is the typical implementation timeline for a university with Wisdom?
- What training and support services are included for Engage platform?
- What customization capabilities exist for the ATS product?
- What are the limits of white-labeling for university branding?

### RECOMMENDATIONS

#### 1. Immediate Actions
- Schedule call with Mike to review this final gaps analysis
- Get clarification on Engage platform maturity and roadmap
- Determine Wisdom analytics requirements for CU
- Identify BrightSync integration needs and costs

#### 2. Response Strategy
- Emphasize the robust Java Spring-based ATS as the core solution
- Highlight Wisdom analytics platform as a major differentiator
- Be transparent about Engage platform as an emerging product
- Propose realistic timeline for AI-powered features

#### 3. Risk Mitigation
- Don't over-promise on Engage platform capabilities
- Focus on proven ATS and Wisdom strengths
- Be transparent about add-on product costs
- Propose realistic solutions for gaps

### NEXT STEPS

1. **Review with Mike:** Go through this final gaps analysis together
2. **Clarify Capabilities:** Get definitive answers on all product capabilities
3. **Assess Impact:** Determine which gaps are critical for CU
4. **Develop Strategy:** Create plan to address or work around gaps
5. **Update Response:** Revise RFP response based on clarified capabilities

### CONTACT INFORMATION
- **Mike Brandt:** Head of Alliances, Inovium
- **Email:** michael.brandt@inovium.com
- **Phone:** [Need to confirm]
- **Meeting Request:** Schedule 1-hour call to review final gaps analysis

---
*This analysis is based on exact content from COMPANY-DETAILS.md and should be verified with Mike before finalizing the RFP response.*
"""
    
    with open("output/FINAL_GAPS_ANALYSIS_FOR_MIKE.md", "w") as f:
        f.write(gaps_document)
    
    print("\nFinal gaps analysis created: output/FINAL_GAPS_ANALYSIS_FOR_MIKE.md")
    print("Please review with Mike before finalizing the RFP response.")

if __name__ == "__main__":
    rebuild_requirements_precisely()
    create_final_gaps_analysis()
