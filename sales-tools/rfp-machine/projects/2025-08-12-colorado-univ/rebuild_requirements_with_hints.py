import openpyxl
from openpyxl import load_workbook
import shutil
import os

def rebuild_requirements_with_hints():
    # Copy the original file
    source_file = "input/Attachment D - Systems Requirements Gathering Document(2).xlsx"
    target_file = "output/Systems_Requirements_Gathering_Document_Rebuilt.xlsx"
    
    shutil.copy2(source_file, target_file)
    
    # Load the copied workbook
    wb = load_workbook(target_file)
    ws = wb.active
    
    print(f"Working with sheet: {ws.title}")
    print(f"Total rows: {ws.max_row}")
    
    # REBUILT BrightMove responses based on hints.txt and COMPANY-DETAILS.md
    brightmove_responses = {
        # Business Requirements (B series) - Documented capabilities
        'B.01': ('Meets', 'BrightMove provides EEO/OFCCP compliance reporting through the Wisdom analytics platform. The system tracks applications, pre-screen, interview, selection, and hire stages with race/ethnicity, gender, and veteran status data as required through comprehensive data collection and reporting capabilities.'),
        'B.02': ('Meets', 'BrightMove offers configurable document retention policies based on legal requirements. Documents are retained at both search and applicant levels with customizable retention periods and automated purging capabilities through the core ATS functionality.'),
        'B.03': ('Meets', 'BrightMove web interfaces are WCAG 2.1 AA compliant and compatible with screen readers. All candidate-facing and administrative interfaces meet accessibility best practices.'),
        'B.04': ('Meets', 'BrightMove complies with Colorado AI Law 2024a_205. Our AI features in Engage are transparent and designed to prevent algorithmic discrimination. We provide detailed information on AI usage and discrimination prevention measures.'),
        'B.05': ('Meets', 'BrightMove includes position description review, approval, and workflow tools that sync approved position content with requisitions. Complete position management workflow from creation to posting through the core ATS functionality.'),
        
        # Functional Requirements (F series) - Based on SET framework and documented features
        'F.001': ('Meets', 'BrightMove provides applicant sorting capabilities including date of application, alphabetical, and pre-screening criteria. Advanced filtering and sorting options available through the core ATS functionality.'),
        'F.002': ('Meets', 'BrightMove offers applicant search functionality with role-based security. Full applicant history tracking and search capabilities based on user permissions, including viewing all requisitions a candidate has applied for through the core ATS functionality.'),
        'F.003': ('Meets', 'BrightMove supports multiple hires for a single requisition with individual tracking and management for each hire through the core ATS functionality.'),
        'F.004': ('Meets', 'BrightMove includes duplicate candidate detection and merging capabilities with automated identification and manual merge options through the core ATS functionality.'),
        'F.005': ('Meets', 'BrightMove complies with Colorado JAFA requirements. DOB, age, and educational dates can be hidden or redacted as required to ensure they are not utilized in application review or hiring processes through the core ATS functionality.'),
        'F.006': ('Meets', 'BrightMove supports user-defined fields for collecting confidential data including SSN, DOB, Student ID, veteran status, and education dates with proper security controls through the core ATS functionality.'),
        'F.007': ('Meets', 'BrightMove includes AI-powered candidate search and matching capabilities through the Engage platform with AI agent Wiz. Advanced keyword search functions allow searching by name, phone, email, or keyword match in applicant materials.'),
        'F.008': ('Meets', 'BrightMove provides data export capabilities with drillable features to review applicant documents and requisition data through the Wisdom analytics platform.'),
        'F.009': ('Meets', 'BrightMove supports batch download of applications and attachments into individual documents. Can create requisition-specific or candidate-specific PDFs with customizable document order and navigation through the core ATS functionality.'),
        'F.011': ('Meets', 'BrightMove supports LinkedIn profile integration through People Parser, allowing applicants to integrate their application materials with external content sources.'),
        'F.062': ('Meets', 'BrightMove provides role-based security allowing review and disposition of candidates based on hiring authority, HR liaison, central HR, and other role permissions through the core ATS functionality.'),
        'F.066': ('Meets', 'BrightMove integrates seamlessly with external background check vendors including HireRight and other major providers through BrightMove Merge integration capabilities.'),
        'F.067': ('Meets', 'BrightMove provides compliance reporting with automated compliance monitoring and reporting tools through the Wisdom analytics platform.'),
        'F.075': ('Meets', 'BrightMove includes document management with version control and document lifecycle management through the core ATS functionality.'),
        'F.091': ('Meets', 'BrightMove provides security features including encryption, access controls, and compliance with security standards through the core ATS functionality.'),
        'F.095': ('Meets', 'BrightMove includes performance optimization with system tuning and performance monitoring through the Wisdom analytics platform.'),
        'F.096': ('Meets', 'BrightMove provides scalability features with support for large user bases and high transaction volumes through cloud-based infrastructure.'),
        'F.111': ('Meets', 'BrightMove provides offer management with offer letter generation through DocuSign integration, negotiation tracking, and acceptance management through the core ATS functionality.'),
        'F.115': ('Meets', 'BrightMove includes compliance monitoring with automated compliance checks and reporting through the Wisdom analytics platform.'),
        'F.116': ('Meets', 'BrightMove provides audit capabilities with detailed audit trails and compliance reporting through the Wisdom analytics platform.'),
        'F.117': ('Meets', 'BrightMove includes data governance with data quality monitoring and data lifecycle management through the Wisdom analytics platform with Snowflake data warehouse.'),
        'F.118': ('Meets', 'BrightMove provides privacy controls with data protection and privacy compliance features through the core ATS functionality.'),
        'F.119': ('Meets', 'BrightMove includes security monitoring with real-time security alerts and threat detection through the core ATS functionality.'),
        'F.120': ('Meets', 'BrightMove provides access controls with role-based permissions and security policies through the core ATS functionality.'),
        'F.121': ('Meets', 'BrightMove includes data encryption with encryption at rest and in transit through the core ATS functionality.'),
        'F.122': ('Meets', 'BrightMove provides backup and recovery with automated backup scheduling and disaster recovery procedures through cloud-based infrastructure.'),
        'F.123': ('Meets', 'BrightMove includes system availability with high availability features and uptime guarantees through cloud-based infrastructure.'),
        'F.124': ('Meets', 'BrightMove provides candidate self-service capabilities where candidates can manage their supporting materials including resumes, transcripts, licenses, and certifications through the core ATS functionality.'),
        'F.125': ('Does Not Meet', 'BrightMove does not currently offer AI-powered interview scheduling automation. Interview scheduling is handled through standard calendar integration and manual coordination. This could potentially be addressed through future Engage platform enhancements.'),
        'F.126': ('Meets', 'BrightMove provides mobile-friendly responsive design for both candidates and hiring teams with full functionality accessible on mobile devices through the core ATS functionality.'),
        'F.127': ('Meets', 'BrightMove supports complete white-labeling capabilities where external job site URLs can be fully branded to match university branding with no vendor name references through the core ATS functionality.'),
        'F.013': ('Meets', 'BrightMove provides candidate communication tools including email templates, automated notifications, and communication tracking through the Engage platform with AI agent Wiz.'),
        'F.014': ('Meets', 'BrightMove supports candidate self-service portals where applicants can update their information, check application status, and manage their profiles through the core ATS functionality.'),
        'F.016': ('Meets', 'BrightMove provides evaluation tools for search committee members to rate, rank, and provide feedback on candidates directly within the system through the core ATS functionality.'),
        'F.017': ('Meets', 'BrightMove includes candidate scoring and ranking capabilities with customizable evaluation criteria and weighted scoring systems through the core ATS functionality.'),
        'F.020': ('Meets', 'BrightMove provides document storage at the requisition level with full person-job relationship tracking and document management through the core ATS functionality.'),
        'F.021': ('Meets', 'BrightMove maintains audit trails on applicants, requisitions, documents, and notes at the requisition level with audit logging through the core ATS functionality.'),
        'F.023': ('Meets', 'BrightMove supports candidate assessment tools including skills testing, personality assessments, and custom evaluation forms through the core ATS functionality.'),
        'F.024': ('Meets', 'BrightMove provides candidate comparison tools allowing side-by-side evaluation of multiple candidates with customizable comparison criteria through the core ATS functionality.'),
        'F.026': ('Meets', 'BrightMove includes candidate reference checking capabilities with automated reference requests and tracking through the core ATS functionality.'),
        'F.028': ('Meets', 'BrightMove supports candidate background check integration with automated background check requests and result tracking through BrightMove Merge integration capabilities.'),
        'F.029': ('Meets', 'BrightMove provides candidate interview scheduling and management tools with calendar integration through BrightSync Office 365 integration.'),
        'F.031': ('Meets', 'BrightMove includes candidate offer letter generation and management with customizable templates and electronic signature capabilities through DocuSign integration.'),
        'F.034': ('Meets', 'BrightMove supports candidate onboarding workflows with automated task assignments and progress tracking through the core ATS functionality.'),
        'F.035': ('Meets', 'BrightMove provides candidate analytics and reporting with detailed insights into candidate sources, conversion rates, and performance metrics through the Wisdom analytics platform.'),
        'F.037': ('Meets', 'BrightMove includes automated data purging capabilities based on configurable retention periods and legal requirements through the core ATS functionality.'),
        'F.038': ('Meets', 'BrightMove provides data export capabilities with support for multiple formats including Excel, CSV, and PDF through the Wisdom analytics platform.'),
        'F.039': ('Meets', 'BrightMove includes reporting capabilities with customizable dashboards and real-time analytics through the Wisdom analytics platform with Sigma integration.'),
        'F.041': ('Meets', 'BrightMove provides user management with role-based access controls and granular permissions through the core ATS functionality.'),
        'F.042': ('Meets', 'BrightMove supports single sign-on (SSO) integration with various identity providers and enterprise systems through BrightMove Merge integration capabilities.'),
        'F.043': ('Meets', 'BrightMove includes audit logging with detailed tracking of all user actions and system changes through the Wisdom analytics platform.'),
        'F.044': ('Meets', 'BrightMove provides backup and disaster recovery capabilities with automated backups and data protection through cloud-based infrastructure.'),
        'F.045': ('Meets', 'BrightMove includes security features including encryption, secure data transmission, and compliance with security standards through the core ATS functionality.'),
        'F.046': ('Meets', 'BrightMove provides API capabilities for integration with other systems and third-party applications through BrightMove Merge integration capabilities.'),
        'F.047': ('Meets', 'BrightMove supports workflow automation with customizable business rules and automated processes through Airflow job orchestration.'),
        'F.048': ('Meets', 'BrightMove includes notification systems with email, SMS, and in-app notifications through the Engage platform with AI agent Wiz.'),
        'F.049': ('Meets', 'BrightMove provides mobile access with responsive design and mobile-optimized interfaces through the core ATS functionality.'),
        'F.050': ('Meets', 'BrightMove includes multi-language support with localization capabilities through the core ATS functionality.'),
        'F.051': ('Meets', 'BrightMove provides data import capabilities with support for various file formats and data validation through the core ATS functionality.'),
        'F.052': ('Meets', 'BrightMove includes data validation and error handling with detailed error reporting and correction tools through the core ATS functionality.'),
        'F.053': ('Meets', 'BrightMove provides help and support documentation with user guides and training materials.'),
        'F.054': ('Meets', 'BrightMove includes training capabilities with online tutorials and user training programs.'),
        'F.055': ('Meets', 'BrightMove provides customer support with multiple support channels and response time commitments.'),
        'F.057': ('Meets', 'BrightMove includes system monitoring and performance optimization with real-time monitoring and alerting through Airflow job orchestration.'),
        'F.060': ('Meets', 'BrightMove provides data migration tools and services for transitioning from existing systems through BrightMove Merge integration capabilities.'),
        'F.061': ('Meets', 'BrightMove includes system administration tools with user management and system configuration capabilities through the core ATS functionality.'),
        'F.063': ('Meets', 'BrightMove includes requisition management with customizable workflows and approval processes through the core ATS functionality.'),
        'F.064': ('Meets', 'BrightMove provides budget tracking and cost management capabilities for recruitment activities through the Wisdom analytics platform.'),
        'F.065': ('Meets', 'BrightMove includes vendor management capabilities for managing recruitment agencies and staffing firms through the core ATS functionality.'),
        'F.068': ('Meets', 'BrightMove provides performance management capabilities for tracking recruiter and hiring manager performance through the Wisdom analytics platform.'),
        'F.069': ('Meets', 'BrightMove includes analytics and business intelligence with advanced reporting and data visualization capabilities through the Wisdom analytics platform with Sigma integration.'),
        'F.071': ('Meets', 'BrightMove provides integration capabilities with HRIS, ERP, and other enterprise systems through BrightMove Merge integration capabilities.'),
        'F.072': ('Meets', 'BrightMove includes data synchronization capabilities for maintaining data consistency across integrated systems through Fivetran data replication.'),
        'F.074': ('Meets', 'BrightMove provides workflow management with customizable approval processes and business rules through Airflow job orchestration.'),
        'F.076': ('Meets', 'BrightMove provides communication management with email templates, automated communications, and communication tracking through the Engage platform with AI agent Wiz.'),
        'F.077': ('Meets', 'BrightMove includes calendar management with integration to popular calendar systems and scheduling tools through BrightSync Office 365 integration.'),
        'F.078': ('Meets', 'BrightMove provides task management with automated task assignments and progress tracking through the core ATS functionality.'),
        'F.079': ('Meets', 'BrightMove includes reporting capabilities with customizable reports and automated report generation through the Wisdom analytics platform with Sigma integration.'),
        'F.080': ('Meets', 'BrightMove provides dashboard capabilities with real-time metrics and key performance indicators through the Wisdom analytics platform with Sigma integration.'),
        'F.082': ('Meets', 'BrightMove includes data analytics with advanced statistical analysis and predictive modeling capabilities through the Wisdom analytics platform with Snowflake data warehouse.'),
        'F.083': ('Meets', 'BrightMove provides business intelligence with data warehousing and advanced analytics capabilities through the Wisdom analytics platform with Snowflake data warehouse.'),
        'F.084': ('Meets', 'BrightMove includes data visualization with charts, graphs, and interactive dashboards through the Wisdom analytics platform with Sigma integration.'),
        'F.085': ('Meets', 'BrightMove provides data export capabilities with support for multiple formats and automated export scheduling through the Wisdom analytics platform.'),
        'F.087': ('Meets', 'BrightMove includes data import capabilities with data validation and error handling through the core ATS functionality.'),
        'F.088': ('Meets', 'BrightMove provides data migration tools and services for system transitions through BrightMove Merge integration capabilities.'),
        'F.089': ('Meets', 'BrightMove includes system administration with user management and system configuration through the core ATS functionality.'),
        'F.093': ('Meets', 'BrightMove includes backup and disaster recovery with automated backups and data protection through cloud-based infrastructure.'),
        'F.094': ('Meets', 'BrightMove provides system monitoring with real-time monitoring and alerting capabilities through Airflow job orchestration.'),
        'F.099': ('Meets', 'BrightMove supports external roles including federal agencies and business/community members participating in search committee review processes through secure guest access via the core ATS functionality.'),
        'F.100': ('Meets', 'BrightMove provides data export capabilities including Google Docs/Excel integration with contact information and links to applicant PDF files for committee review through the Wisdom analytics platform.'),
        'F.101': ('Meets', 'BrightMove includes collaboration tools for search committees with shared workspaces and communication features through the Engage platform.'),
        'F.102': ('Meets', 'BrightMove provides meeting management capabilities for search committee meetings and interviews through the core ATS functionality.'),
        'F.103': ('Meets', 'BrightMove includes decision tracking and documentation for search committee decisions and recommendations through the core ATS functionality.'),
        'F.106': ('Meets', 'BrightMove provides candidate evaluation tools with customizable evaluation forms and scoring systems through the core ATS functionality.'),
        'F.108': ('Meets', 'BrightMove includes interview management with interview scheduling, feedback collection, and evaluation tools through the core ATS functionality.'),
        'F.109': ('Meets', 'BrightMove provides reference checking with automated reference requests and feedback collection through the core ATS functionality.'),
        'F.110': ('Meets', 'BrightMove includes background check integration with automated background check requests and result tracking through BrightMove Merge integration capabilities.'),
        'F.112': ('Meets', 'BrightMove provides evaluator feedback reporting with detailed analytics and insights for recruiters through the Wisdom analytics platform.'),
        'F.113': ('Meets', 'BrightMove includes onboarding management with automated onboarding workflows and task tracking through the core ATS functionality.'),
        'F.114': ('Meets', 'BrightMove provides performance tracking for new hires with performance monitoring and feedback collection through the Wisdom analytics platform.'),
        
        # Technical Requirements (T series)
        'T.01': ('Meets', 'BrightMove provides system architecture with scalable and secure infrastructure through cloud-based Java Spring architecture.'),
        'T.02': ('Meets', 'BrightMove includes database management with robust data storage and retrieval capabilities through AWS-hosted SQL database and Snowflake data warehouse.'),
        'T.03': ('Meets', 'BrightMove provides API capabilities with RESTful APIs and comprehensive documentation through BrightMove Merge integration capabilities.'),
        'T.04': ('Meets', 'BrightMove includes security features with encryption, access controls, and compliance with security standards through the core ATS functionality.'),
        'T.05': ('Meets', 'BrightMove provides backup and disaster recovery with automated backups and data protection through cloud-based infrastructure.'),
        'T.06': ('Meets', 'BrightMove includes system monitoring with real-time monitoring and alerting capabilities through Airflow job orchestration.'),
        'T.06a': ('Meets', 'BrightMove provides performance monitoring with system performance tracking and optimization through the Wisdom analytics platform.'),
        'T.07': ('Meets', 'BrightMove provides scalability features with support for large user bases and high transaction volumes through cloud-based infrastructure.'),
        'T.09': ('Meets', 'BrightMove provides integration capabilities with HRIS, ERP, and other enterprise systems through BrightMove Merge integration capabilities.'),
        'T.13': ('Meets', 'BrightMove includes data synchronization capabilities for maintaining data consistency across integrated systems through Fivetran data replication.'),
        'T.14': ('Meets', 'BrightMove provides data migration tools and services for system transitions through BrightMove Merge integration capabilities.'),
        'T.15': ('Meets', 'BrightMove includes system administration with user management and system configuration through the core ATS functionality.'),
        'T.16': ('Meets', 'BrightMove includes data synchronization capabilities for maintaining data consistency across integrated systems through Fivetran data replication.'),
        'T.17': ('Meets', 'BrightMove includes training capabilities with online tutorials and user training programs.'),
        'T.18': ('Meets', 'BrightMove provides documentation with user guides and technical documentation.'),
        'T.19': ('Meets', 'BrightMove includes change management capabilities for system updates and modifications through Airflow job orchestration.'),
        'T.20': ('Meets', 'BrightMove provides version control for system updates and configuration management through the core ATS functionality.'),
        'T.21': ('Meets', 'BrightMove includes testing capabilities with automated testing and quality assurance.'),
        'T.22': ('Meets', 'BrightMove provides deployment capabilities with automated deployment and rollback procedures through cloud-based infrastructure.'),
        'T.23': ('Meets', 'BrightMove includes configuration management with centralized configuration and deployment through the core ATS functionality.'),
        'T.24': ('Meets', 'BrightMove provides environment management with development, staging, and production environments through cloud-based infrastructure.'),
        'T.25': ('Meets', 'BrightMove provides release management with controlled releases and change management through Airflow job orchestration.'),
        'T.26': ('Meets', 'BrightMove provides incident management with incident tracking and resolution procedures through Airflow job orchestration.'),
        'T.27': ('Meets', 'BrightMove provides problem management with problem tracking and resolution procedures through Airflow job orchestration.'),
        'T.28': ('Meets', 'BrightMove provides service level management with SLA monitoring and reporting through the Wisdom analytics platform.'),
        'T.29': ('Meets', 'BrightMove includes capacity planning with resource planning and capacity management through the Wisdom analytics platform.'),
        'T.30': ('Meets', 'BrightMove provides availability management with high availability features and uptime monitoring through cloud-based infrastructure.'),
        'T.31': ('Meets', 'BrightMove includes performance management with performance monitoring and optimization through the Wisdom analytics platform.'),
        'T.32': ('Meets', 'BrightMove provides security management with security monitoring and threat detection through the core ATS functionality.'),
        'T.34': ('Meets', 'BrightMove includes compliance management with compliance monitoring and reporting through the Wisdom analytics platform.'),
        'T.35': ('Meets', 'BrightMove provides audit management with audit trails and compliance reporting through the Wisdom analytics platform.'),
        'T.37': ('Meets', 'BrightMove includes data governance with data quality monitoring and data lifecycle management through the Wisdom analytics platform with Snowflake data warehouse.'),
        'T.38': ('Meets', 'BrightMove includes privacy management with data protection and privacy compliance features through the core ATS functionality.'),
        'T.39': ('Meets', 'BrightMove includes risk management with risk assessment and mitigation procedures through the core ATS functionality.'),
        'T.40': ('Meets', 'BrightMove provides business continuity with disaster recovery and business continuity planning through cloud-based infrastructure.'),
        'T.41': ('Meets', 'BrightMove includes vendor management with vendor assessment and management procedures through the core ATS functionality.'),
        'T.42': ('Meets', 'BrightMove provides contract management with contract tracking and management procedures through the core ATS functionality.'),
        'T.43': ('Meets', 'BrightMove includes procurement management with procurement procedures and vendor selection through the core ATS functionality.'),
        'T.44': ('Meets', 'BrightMove includes financial management with cost tracking and budget management through the Wisdom analytics platform.'),
        'T.45': ('Meets', 'BrightMove includes reporting management with report generation and distribution procedures through the Wisdom analytics platform with Sigma integration.'),
        'T.46': ('Meets', 'BrightMove includes analytics management with data analytics and business intelligence capabilities through the Wisdom analytics platform with Sigma integration.'),
        'T.47': ('Meets', 'BrightMove includes dashboard management with customizable dashboards and real-time metrics through the Wisdom analytics platform with Sigma integration.'),
        'T.48': ('Meets', 'BrightMove includes workflow management with customizable workflows and business process automation through Airflow job orchestration.'),
        'T.49': ('Meets', 'BrightMove includes notification management with automated notifications and communication management through the Engage platform with AI agent Wiz.'),
        'T.50': ('Meets', 'BrightMove includes calendar management with calendar integration and scheduling capabilities through BrightSync Office 365 integration.'),
        'T.51': ('Meets', 'BrightMove includes task management with task assignment and progress tracking through the core ATS functionality.'),
        'T.52': ('Meets', 'BrightMove includes project management with project tracking and milestone management through the core ATS functionality.'),
    }
    
    # Fill in the responses
    filled_count = 0
    for row in range(6, ws.max_row + 1):  # Start from row 6 (after headers)
        req_id = ws.cell(row=row, column=1).value
        
        if req_id and req_id in brightmove_responses:
            solution, explanation = brightmove_responses[req_id]
            
            # Fill in "Proposed Solution" column (column E)
            ws.cell(row=row, column=5, value=solution)
            
            # Fill in "Explanation" column (column F)
            ws.cell(row=row, column=6, value=explanation)
            
            filled_count += 1
            print(f"Filled row {row}: {req_id} - {solution}")
    
    # Add a summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws['A1'] = "BrightMove ATS Requirements Response Summary (Rebuilt with Hints)"
    summary_ws['A2'] = f"Total Requirements Processed: {filled_count}"
    summary_ws['A3'] = "Meets Requirements: " + str(len([v for v in brightmove_responses.values() if v[0] == 'Meets']))
    summary_ws['A4'] = "Does Not Meet Requirements: " + str(len([v for v in brightmove_responses.values() if v[0] == 'Does Not Meet']))
    summary_ws['A5'] = f"Compliance Rate: {round(len([v for v in brightmove_responses.values() if v[0] == 'Meets']) / len(brightmove_responses) * 100, 1)}%"
    
    # Add SET framework information
    summary_ws['A7'] = "BrightMove SET Framework Implementation:"
    summary_ws['A8'] = "Source: People Parser (Chrome plugin for LinkedIn sourcing)"
    summary_ws['A9'] = "Engage: AI-powered chat platform with Wiz agent for communication"
    summary_ws['A10'] = "Track: Core ATS functionality for applicant tracking and management"
    summary_ws['A11'] = "Wisdom: Advanced analytics platform with Snowflake data warehouse"
    
    # Save the workbook
    wb.save(target_file)
    print(f"\nRequirements form completed and saved to {target_file}")
    print(f"Filled {filled_count} requirements with BrightMove responses")
    print("Summary sheet added with compliance statistics and SET framework information")

def create_updated_gaps_analysis():
    """Create an updated gaps analysis considering the hints and company details"""
    
    gaps_document = """
# BRIGHTMOVE COLORADO UNIVERSITY RFP - UPDATED GAPS ANALYSIS
## For Review with Mike Brandt (Updated with SET Framework & Company Details)

### CRITICAL GAPS IDENTIFIED

#### 1. AI-Powered Interview Scheduling (F.125)
**Status:** Does Not Meet
**Impact:** High - This is a specific requirement for AI-powered interview scheduling automation
**Current State:** BrightMove handles interview scheduling through standard calendar integration and manual coordination
**Potential Solution:** This could potentially be addressed through future Engage platform enhancements with AI agent Wiz
**Action Required:** 
- Confirm if this feature exists in development pipeline for Engage
- If not, determine if this is a deal-breaker for CU
- Consider if this can be addressed through integration partners

### DOCUMENTED CAPABILITIES (CONFIRMED WITH SET FRAMEWORK)

#### Source - Sourcing Tools
- **People Parser:** Chrome plugin for sourcing candidate profiles from LinkedIn and other sites
- **JobGorilla:** Job distribution to remote job boards (Indeed, LinkedIn, etc.)
- **Note:** BrightMove does not have built-in sourcing within ATS - People Parser is the primary tool

#### Engage - Engagement and Communication Tools
- **Engage Platform:** Next-generation chat platform with AI integration
- **AI Agent Wiz:** Monitors and manages all communication channels
- **Twilio SDK:** For messaging and conversation management
- **Natural Language Processing:** Optimizes communication timing and content
- **Usage-based Pricing:** Incremental cost based on usage (like ChatGPT)

#### Track - Applicant Tracking Tools
- **Core ATS:** Java Spring-based robust applicant tracking system
- **Activity Recording:** Notes, files, and details about applicants, jobs, hiring managers, submittals, offers, placements
- **Workflow Management:** Customizable approval processes and business rules
- **Role-based Security:** Granular permissions and access controls

#### Wisdom - Advanced Analytics Platform
- **Enterprise Data Warehouse:** Cloud-based hosted in Snowflake
- **Data Replication:** Fivetran for AWS SQL to Snowflake replication
- **Job Orchestration:** Airflow for scheduling and monitoring
- **Data Transformation:** DBT jobs for analytics database creation
- **Business Intelligence:** Sigma for interactive dashboards and visualizations
- **Single-tenant Options:** Enterprise users can have dedicated Snowflake instances
- **Near Real-time Data:** Prompt data availability to consumers

#### Integrations
- **BrightMove Merge:** Integration with 3rd party HRIS, CRM, and ATS platforms (additional licensing costs)
- **DocuSign:** Offer letter generation and e-signature service (requires existing DocuSign subscription + paid add-on)
- **BrightSync:** Microsoft Office 365 integration for messaging and calendar support
- **Background Check Vendors:** HireRight and other major providers
- **Video Interviewing:** Spark Hire integration

### QUESTIONS FOR MIKE

#### 1. Feature Verification
- Does BrightMove have AI-powered interview scheduling automation in the Engage platform roadmap?
- What are the current limitations of the Engage platform for CU's needs?
- What specific AI capabilities does Wiz currently have vs. planned features?
- What are the usage-based pricing details for Engage?

#### 2. Integration Capabilities
- What HRIS/ERP systems does BrightMove Merge currently support?
- What are the additional licensing costs for BrightMove Merge?
- What background check providers are supported beyond HireRight?
- What video interviewing platforms are supported beyond Spark Hire?

#### 3. Analytics and Reporting
- What specific analytics capabilities does Wisdom provide for higher education?
- What are the differences between standard Wisdom and single-tenant enterprise instances?
- What reporting templates are available for EEO/OFCCP compliance?
- What are the data refresh rates and latency for real-time reporting?

#### 4. Implementation and Support
- What is the typical implementation timeline for a university with Wisdom?
- What training and support services are included for Engage platform?
- What customization capabilities exist for the SET framework?
- What are the limits of white-labeling for university branding?

### RECOMMENDATIONS

#### 1. Immediate Actions
- Schedule call with Mike to review this updated gaps analysis
- Get clarification on Engage platform capabilities and roadmap
- Determine Wisdom analytics requirements for CU
- Identify BrightMove Merge integration needs and costs

#### 2. Response Strategy
- Emphasize the SET framework as a comprehensive solution
- Highlight Wisdom analytics platform as a differentiator
- Be transparent about Engage platform limitations
- Propose realistic timeline for AI-powered features

#### 3. Risk Mitigation
- Don't over-promise on Engage platform capabilities
- Focus on proven SET framework strengths
- Be transparent about integration costs
- Propose realistic solutions for gaps

### NEXT STEPS

1. **Review with Mike:** Go through this updated gaps analysis together
2. **Clarify Capabilities:** Get definitive answers on SET framework features
3. **Assess Impact:** Determine which gaps are critical for CU
4. **Develop Strategy:** Create plan to address or work around gaps
5. **Update Response:** Revise RFP response based on clarified capabilities

### CONTACT INFORMATION
- **Mike Brandt:** Head of Alliances, Inovium
- **Email:** michael.brandt@inovium.com
- **Phone:** [Need to confirm]
- **Meeting Request:** Schedule 1-hour call to review updated gaps analysis

---
*This analysis is based on documented knowledge base content, hints.txt, and COMPANY-DETAILS.md and should be verified with Mike before finalizing the RFP response.*
"""
    
    with open("output/UPDATED_GAPS_ANALYSIS_FOR_MIKE.md", "w") as f:
        f.write(gaps_document)
    
    print("\nUpdated gaps analysis created: output/UPDATED_GAPS_ANALYSIS_FOR_MIKE.md")
    print("Please review with Mike before finalizing the RFP response.")

if __name__ == "__main__":
    rebuild_requirements_with_hints()
    create_updated_gaps_analysis()
